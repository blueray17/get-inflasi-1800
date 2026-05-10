import subprocess, sys, importlib

def ensure_package(package, import_name=None):
    try:
        importlib.import_module(import_name or package)
    except ModuleNotFoundError:
        subprocess.check_call([sys.executable, "-m", "pip", "install", package, "-q"])

ensure_package("gspread")
ensure_package("google-auth", "google.auth")
ensure_package("openpyxl")

import streamlit as st
import pandas as pd
import gspread
from google.oauth2 import service_account
import io, re, json, requests

st.set_page_config(page_title="Generator Data Inflasi", page_icon="📊", layout="centered")

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Plus+Jakarta+Sans:wght@400;500;600;700&family=JetBrains+Mono:wght@400;600&display=swap');
html,body,[class*="css"]{font-family:'Plus Jakarta Sans',sans-serif;}
.stApp{background:linear-gradient(135deg,#0f172a 0%,#1e293b 50%,#0f172a 100%);min-height:100vh;}
.main-header{text-align:center;padding:2.5rem 0 1.5rem;}
.main-header h1{font-size:2rem;font-weight:700;color:#f1f5f9;margin:0;letter-spacing:-0.5px;}
.main-header p{color:#64748b;font-size:0.9rem;margin-top:.4rem;}
.badge{display:inline-block;background:linear-gradient(90deg,#3b82f6,#6366f1);color:#fff;
  font-size:.7rem;font-weight:600;padding:3px 10px;border-radius:20px;letter-spacing:1px;
  text-transform:uppercase;margin-bottom:.8rem;}
.card{background:rgba(30,41,59,.8);border:1px solid rgba(99,102,241,.2);
  border-radius:16px;padding:1.5rem;margin-bottom:1.2rem;}
.card-title{font-size:.75rem;font-weight:600;color:#6366f1;text-transform:uppercase;
  letter-spacing:1.5px;margin-bottom:1rem;}
label{color:#94a3b8 !important;font-size:.82rem !important;font-weight:500 !important;}
.stTextInput>div>div>input,.stNumberInput>div>div>input{
  background:rgba(15,23,42,.8)!important;border:1px solid rgba(99,102,241,.3)!important;
  border-radius:8px!important;color:#f1f5f9!important;font-family:'JetBrains Mono',monospace!important;}
.stSelectbox>div>div{background:rgba(15,23,42,.8)!important;
  border:1px solid rgba(99,102,241,.3)!important;border-radius:8px!important;color:#f1f5f9!important;}
.stButton>button{width:100%;background:linear-gradient(135deg,#3b82f6,#6366f1)!important;
  color:#fff!important;border:none!important;border-radius:10px!important;
  padding:.75rem 2rem!important;font-size:.95rem!important;font-weight:600!important;}
.stDownloadButton>button{width:100%;background:linear-gradient(135deg,#10b981,#059669)!important;
  color:#fff!important;border:none!important;border-radius:10px!important;
  padding:.75rem 2rem!important;font-size:.95rem!important;font-weight:600!important;}
.info-box{background:rgba(59,130,246,.1);border-left:3px solid #3b82f6;
  border-radius:0 8px 8px 0;padding:.8rem 1rem;margin:.8rem 0;color:#93c5fd;font-size:.83rem;}
.success-box{background:rgba(16,185,129,.1);border-left:3px solid #10b981;
  border-radius:0 8px 8px 0;padding:.8rem 1rem;margin:.8rem 0;color:#6ee7b7;font-size:.83rem;}
.error-box{background:rgba(239,68,68,.1);border-left:3px solid #ef4444;
  border-radius:0 8px 8px 0;padding:.8rem 1rem;margin:.8rem 0;color:#fca5a5;font-size:.83rem;}
.warn-box{background:rgba(245,158,11,.1);border-left:3px solid #f59e0b;
  border-radius:0 8px 8px 0;padding:.8rem 1rem;margin:.8rem 0;color:#fcd34d;font-size:.83rem;}
.stat-row{display:flex;gap:1rem;margin-top:.8rem;}
.stat-item{flex:1;background:rgba(15,23,42,.6);border-radius:10px;padding:.8rem;text-align:center;}
.stat-value{font-size:1.4rem;font-weight:700;color:#6366f1;font-family:'JetBrains Mono',monospace;}
.stat-label{font-size:.7rem;color:#64748b;margin-top:2px;}
.footer{text-align:center;color:#334155;font-size:.75rem;padding:2rem 0 1rem;}
</style>
""", unsafe_allow_html=True)

# ── Constants ─────────────────────────────────────────────────────────────────
SPREADSHEET_ID  = "15HbcEJwdK9TUo8Wpkgqnfveyp67RLK4B"
SPREADSHEET_URL = f"https://docs.google.com/spreadsheets/d/{SPREADSHEET_ID}/edit"

KODE_WILAYAH = {0:"1800", 1:"1804", 2:"1811", 3:"1871", 4:"1872"}
NAMA_WILAYAH = {0:"Provinsi Lampung", 1:"Lampung Timur", 2:"Mesuji", 3:"Bandar Lampung", 4:"Metro"}

BULAN = [("01","Januari"),("02","Februari"),("03","Maret"),("04","April"),
         ("05","Mei"),("06","Juni"),("07","Juli"),("08","Agustus"),
         ("09","September"),("10","Oktober"),("11","November"),("12","Desember")]

def gen_cols():
    c = list("ABCDEFGHIJKLMNOPQRSTUVWXYZ")
    for a in "ABCDEFGHIJKLMNOPQRSTUVWXYZ":
        for b in "ABCDEFGHIJKLMNOPQRSTUVWXYZ":
            c.append(a+b)
    return c

ALL_COLS = gen_cols()

def col_idx(col):
    r = 0
    for ch in col.upper():
        r = r*26 + (ord(ch)-64)
    return r-1

# ── Fetch helpers ─────────────────────────────────────────────────────────────

def download_as_xlsx(spreadsheet_id, api_key):
    """
    Download file sebagai xlsx via Drive API v3.
    Bekerja untuk file Office (.xlsx) maupun Google Sheets native.
    Untuk file Office: gunakan /files/{id}?alt=media (download langsung)
    Untuk Google Sheets: gunakan /files/{id}/export?mimeType=xlsx
    """
    headers = {}

    # Cek tipe file dulu
    meta_url = f"https://www.googleapis.com/drive/v3/files/{spreadsheet_id}?fields=mimeType,name&key={api_key}"
    r = requests.get(meta_url, timeout=20)
    if r.status_code != 200:
        err = r.json().get("error", {})
        raise Exception(f"Drive API error ({r.status_code}): {err.get('message', r.text[:200])}")

    meta = r.json()
    mime = meta.get("mimeType", "")
    name = meta.get("name", "")

    if mime == "application/vnd.google-apps.spreadsheet":
        # Google Sheets native → export
        dl_url = f"https://www.googleapis.com/drive/v3/files/{spreadsheet_id}/export?mimeType=application%2Fvnd.openxmlformats-officedocument.spreadsheetml.sheet&key={api_key}"
    else:
        # Office file (.xlsx, .xls, dll) → download langsung
        dl_url = f"https://www.googleapis.com/drive/v3/files/{spreadsheet_id}?alt=media&key={api_key}"

    r2 = requests.get(dl_url, timeout=60)
    if r2.status_code != 200:
        err = r2.json().get("error", {}) if r2.headers.get("content-type","").startswith("application/json") else {}
        raise Exception(f"Download error ({r2.status_code}): {err.get('message', r2.text[:200])}")

    return io.BytesIO(r2.content), name, mime


def read_sheets_from_xlsx(xlsx_bytes, sheet_indices):
    """Baca sheet dari bytes xlsx menggunakan openpyxl."""
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_bytes, read_only=True, data_only=True)
    sheet_names = wb.sheetnames
    results = {}
    for i in sheet_indices:
        if i >= len(sheet_names):
            results[i] = (None, f"Sheet ke-{i+1} tidak ada")
            continue
        ws = wb[sheet_names[i]]
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append([("" if v is None else str(v)) for v in row])
        if not rows:
            results[i] = (pd.DataFrame(), sheet_names[i])
            continue
        max_col = max(len(r) for r in rows)
        rows = [r + [""]*(max_col-len(r)) for r in rows]
        results[i] = (pd.DataFrame(rows, dtype=str).fillna(""), sheet_names[i])
    wb.close()
    return results


def fetch_via_api_key(spreadsheet_id, api_key):
    """Download seluruh file, kemudian baca 5 sheet pertama."""
    xlsx_bytes, name, mime = download_as_xlsx(spreadsheet_id, api_key)
    sheet_data = read_sheets_from_xlsx(xlsx_bytes, list(range(5)))
    return sheet_data, name, mime


def fetch_via_service_account(credentials_json):
    """Baca via gspread — untuk Google Sheets native + Service Account."""
    scope = [
        "https://spreadsheets.google.com/feeds",
        "https://www.googleapis.com/auth/spreadsheets",
        "https://www.googleapis.com/auth/drive",
    ]
    creds = service_account.Credentials.from_service_account_info(
        json.loads(credentials_json), scopes=scope
    )
    client = gspread.authorize(creds)
    ss = client.open_by_url(SPREADSHEET_URL)
    worksheets = ss.worksheets()
    result = {}
    for i in range(5):
        if i >= len(worksheets):
            result[i] = (None, f"Sheet ke-{i+1} tidak ada")
            continue
        ws = worksheets[i]
        raw = ws.get_all_values()
        if not raw:
            result[i] = (pd.DataFrame(), ws.title)
            continue
        max_col = max(len(r) for r in raw)
        rows = [r + [""]*(max_col-len(r)) for r in raw]
        result[i] = (pd.DataFrame(rows, dtype=str).fillna(""), ws.title)
    return result

# ── UI ────────────────────────────────────────────────────────────────────────
st.markdown("""
<div class="main-header">
  <div class="badge">BPS Provinsi Lampung</div>
  <h1>📊 Generator Data Inflasi</h1>
  <p>Konversi data Google Sheets → file Excel terstruktur</p>
</div>""", unsafe_allow_html=True)

# ── Sidebar ───────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("### ⚙️ Konfigurasi Akses")
    st.markdown("---")

    auth_mode = st.radio(
        "Mode Autentikasi",
        ["🔑 API Key", "🔐 Service Account"],
        index=0,
    )

    api_key = ""
    credentials_json = ""

    if "API Key" in auth_mode:
        st.markdown("""
        <div class="info-box">
        <b>Syarat API Key:</b><br>
        • Google Cloud project sudah ada<br>
        • <b>Google Sheets API</b> + <b>Google Drive API</b> di-enable<br>
        • Spreadsheet: <i>Anyone with the link → Viewer</i>
        </div>
        """, unsafe_allow_html=True)
        api_key = st.text_input("API Key", type="password", placeholder="AIzaSy...")
        if not api_key:
            st.warning("Masukkan API Key untuk melanjutkan")
    else:
        credentials_json = st.text_area(
            "Service Account JSON",
            placeholder='{"type":"service_account",...}',
            height=160,
        )

    st.markdown("---")
    st.markdown("**📌 Kode Wilayah per Sheet**")
    for i in range(5):
        st.markdown(f"""
        <div style="display:flex;justify-content:space-between;padding:4px 0;
        border-bottom:1px solid rgba(99,102,241,.1);font-size:.78rem;">
          <span style="color:#94a3b8">Sheet {i+1} — {NAMA_WILAYAH[i]}</span>
          <span style="color:#6366f1;font-family:monospace;font-weight:600">{KODE_WILAYAH[i]}</span>
        </div>""", unsafe_allow_html=True)

# ── Form ──────────────────────────────────────────────────────────────────────
st.markdown('<div class="card"><div class="card-title">📅 Parameter Waktu</div>', unsafe_allow_html=True)
c1, c2 = st.columns(2)
with c1:
    tahun = st.number_input("Tahun", min_value=2000, max_value=2099, value=2026, step=1)
with c2:
    bulan_opts = [f"{k}. {n}" for k,n in BULAN]
    bulan_sel  = st.selectbox("Bulan", bulan_opts)
    kode_bulan = bulan_sel.split(".")[0].strip()
st.markdown('</div>', unsafe_allow_html=True)

st.markdown('<div class="card"><div class="card-title">📋 Rentang Kolom Data</div>', unsafe_allow_html=True)
st.markdown('<div class="info-box">Kolom A otomatis diambil sebagai Kode. Pilih kolom tambahan yang ingin diambil.</div>', unsafe_allow_html=True)
c3, c4 = st.columns(2)
with c3:
    kolom_awal  = st.selectbox("Kolom Awal",  ALL_COLS, index=ALL_COLS.index("B"))
with c4:
    kolom_akhir = st.selectbox("Kolom Akhir", ALL_COLS, index=ALL_COLS.index("Z"))

idx_awal  = col_idx(kolom_awal)
idx_akhir = col_idx(kolom_akhir)

if idx_awal > idx_akhir:
    st.markdown('<div class="error-box">⚠️ Kolom awal tidak boleh melebihi kolom akhir</div>', unsafe_allow_html=True)
else:
    st.markdown(f'<div class="success-box">✅ Kolom <b>{kolom_awal}</b> s/d <b>{kolom_akhir}</b> ({idx_akhir-idx_awal+1} kolom)</div>', unsafe_allow_html=True)
st.markdown('</div>', unsafe_allow_html=True)

nama_file = f"data_inflasi_{int(tahun)}_{kode_bulan}.xlsx"
st.markdown(f"""
<div class="card">
  <div class="card-title">📁 Output</div>
  <div style="display:flex;align-items:center;gap:.8rem;">
    <div style="font-size:2rem;">📄</div>
    <div>
      <div style="color:#f1f5f9;font-weight:600;font-family:'JetBrains Mono',monospace;">{nama_file}</div>
      <div style="color:#64748b;font-size:.78rem;margin-top:2px;">
        5 sheet → 1 sheet gabungan &nbsp;•&nbsp; A + Tahun + Bulan + Kode_Wilayah + {kolom_awal}:{kolom_akhir}
      </div>
    </div>
  </div>
</div>""", unsafe_allow_html=True)

# ── Generate ──────────────────────────────────────────────────────────────────
st.markdown("<br>", unsafe_allow_html=True)
if st.button("⚡ Generate Excel", use_container_width=True):
    if idx_awal > idx_akhir:
        st.error("Kolom awal harus ≤ kolom akhir!")
        st.stop()

    use_api_key = "API Key" in auth_mode
    if use_api_key and not api_key.strip():
        st.error("Masukkan API Key di sidebar!")
        st.stop()
    if not use_api_key and not credentials_json.strip():
        st.error("Masukkan Service Account JSON di sidebar!")
        st.stop()

    progress = st.progress(0, text="Mengunduh file dari Google Drive...")
    all_dfs, errors = [], []

    try:
        if use_api_key:
            progress.progress(0.1, text="Mendeteksi tipe file...")
            sheet_data, fname, fmime = fetch_via_api_key(SPREADSHEET_ID, api_key.strip())
            st.markdown(f'<div class="info-box">📄 File: <b>{fname}</b> | Tipe: <code>{fmime}</code></div>', unsafe_allow_html=True)
        else:
            progress.progress(0.1, text="Autentikasi Service Account...")
            sheet_data = fetch_via_service_account(credentials_json)

        for i in range(5):
            progress.progress(0.15 + (i+1)*0.15, text=f"Memproses Sheet {i+1} — {NAMA_WILAYAH[i]}...")

            df_raw, sheet_title = sheet_data.get(i, (None, f"Sheet {i+1}"))

            if df_raw is None:
                errors.append(f"Sheet {i+1}: {sheet_title}")
                continue
            if df_raw.empty:
                errors.append(f"Sheet {i+1} ({sheet_title}) kosong, dilewati.")
                continue

            # Perpanjang kolom jika kurang
            while df_raw.shape[1] <= idx_akhir:
                df_raw[df_raw.shape[1]] = ""

            col_a      = df_raw.iloc[:, 0]
            actual_end = min(idx_akhir, df_raw.shape[1]-1)
            range_data = df_raw.iloc[:, idx_awal:actual_end+1].copy()
            range_data.columns = ALL_COLS[idx_awal:actual_end+1]

            result = pd.DataFrame()
            result["Kode"]         = col_a.values
            result["Tahun"]        = str(int(tahun))
            result["Bulan"]        = kode_bulan
            result["Kode_Wilayah"] = KODE_WILAYAH[i]
            for cn in range_data.columns:
                result[cn] = range_data[cn].values

            all_dfs.append(result)

    except Exception as e:
        progress.empty()
        st.markdown(f'<div class="error-box">❌ {e}</div>', unsafe_allow_html=True)
        st.stop()

    progress.progress(0.95, text="Menyimpan Excel...")

    for err in errors:
        st.markdown(f'<div class="warn-box">⚠️ {err}</div>', unsafe_allow_html=True)

    if not all_dfs:
        progress.empty()
        st.error("Tidak ada data berhasil diambil.")
    else:
        final_df = pd.concat(all_dfs, ignore_index=True)

        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine="openpyxl") as writer:
            final_df.to_excel(writer, index=False, sheet_name="Data Inflasi")
            ws_xl = writer.sheets["Data Inflasi"]
            for col in ws_xl.columns:
                w = max((len(str(c.value or "")) for c in col), default=8)
                ws_xl.column_dimensions[col[0].column_letter].width = min(w+2, 40)
        buf.seek(0)
        progress.empty()

        st.markdown(f"""
        <div class="success-box">✅ Berhasil! {len(all_dfs)} sheet diproses.</div>
        <div class="stat-row">
          <div class="stat-item"><div class="stat-value">{len(all_dfs)}</div><div class="stat-label">Sheet</div></div>
          <div class="stat-item"><div class="stat-value">{len(final_df):,}</div><div class="stat-label">Baris</div></div>
          <div class="stat-item"><div class="stat-value">{final_df.shape[1]}</div><div class="stat-label">Kolom</div></div>
        </div>""", unsafe_allow_html=True)

        with st.expander("👁️ Preview 10 baris pertama"):
            st.dataframe(final_df.head(10), use_container_width=True)

        st.markdown("<br>", unsafe_allow_html=True)
        st.download_button(
            label=f"⬇️ Unduh {nama_file}",
            data=buf.getvalue(),
            file_name=nama_file,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

st.markdown('<div class="footer">BPS Provinsi Lampung • Generator Data Inflasi • susenas.my.id</div>',
            unsafe_allow_html=True)
